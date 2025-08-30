# app/routes.py
import os
import io
import json
import csv
import datetime as dt
from collections import Counter
import pandas as pd
from io import BytesIO
from flask import send_file
import matplotlib
matplotlib.use('Agg')  # Set non-interactive backend
import matplotlib.pyplot as plt
import seaborn as sns
import requests  # needed for the proxy endpoint

from flask import (
    Blueprint, render_template, request, redirect, url_for,
    session, send_file, current_app, Response, make_response
)

from .utils import get_db, login_required, admin_required
from .mood_detector import EMOTIONS, advice_for
from .image_generator import build_image_url
from .logger import log_event
from models.user import verify_credentials, ADMIN_USERNAME
from urllib.parse import quote_plus, urlparse

bp = Blueprint("main", __name__)

# Hosts we are willing to proxy images from (prevents open proxy abuse)
ALLOWED_IMAGE_HOSTS = {"image.pollinations.ai"}

# -------------------------------
# Landing (Admin / Users chooser)
# -------------------------------
@bp.get("/")
def entry():
    # If already logged in, go to app
    if session.get("username"):
        return redirect(url_for("main.home"))
    return render_template("entry.html")

# -------------------------------
# Users login (demo users only)
# -------------------------------
@bp.get("/login")
def login_page():
    # Admin is not exposed here; use /admin-login for admin
    return render_template("login.html")

@bp.post("/login")
def login():
    # Normalize username
    username = (request.form.get("username", "")).strip().lower()

    # Accept multiple field names (typed or dropdown pairing UI)
    password = (
        request.form.get("password") or
        request.form.get("password_select") or
        request.form.get("user_password") or
        request.form.get("pw") or
        ""
    )

    # force admin to use the admin login page
    if username == ADMIN_USERNAME:
        log_event("login_admin_on_user_route", user=username)
        return render_template("login.html", error="Admin must use the Admin button.")

    if verify_credentials(username, password):
        session["username"] = username
        log_event("login_success", user=username)
        return redirect(url_for("main.home"))

    log_event("login_fail", user=username)
    return render_template("login.html", error="Invalid credentials")


# -------------------------------
# Admin login (separate page)
# -------------------------------
@bp.get("/admin-login")
def admin_login_page():
    # Only shows a single username ("admin") + password box
    return render_template("admin_login.html", admin_username=ADMIN_USERNAME)

@bp.post("/admin-login")
def admin_login():
    username = (request.form.get("username", "")).strip()
    password = request.form.get("password", "")

    if username != ADMIN_USERNAME:
        log_event("admin_login_wrong_username", user=username)
        return render_template("admin_login.html", admin_username=ADMIN_USERNAME,
                               error="Invalid admin username.")

    if verify_credentials(username, password):
        session["username"] = username
        log_event("admin_login_success", user=username)
        return redirect(url_for("main.home"))

    log_event("admin_login_fail", user=username)
    return render_template("admin_login.html", admin_username=ADMIN_USERNAME,
                           error="Invalid password.")

# -------------------------------
# Logout → back to chooser
# -------------------------------
@bp.get("/logout")
def logout():
    user = session.get("username")
    session.clear()
    log_event("logout", user=user)
    return redirect(url_for("main.entry"))

# -------------------------------
# Main app (home)
# -------------------------------
@bp.get("/home")
@login_required
def home():
    return render_template("home.html", emotions=EMOTIONS, admin_username=ADMIN_USERNAME)

# HTMX endpoint to generate (no full reload)
@bp.post("/generate")
@login_required
def generate():
    emotion = (request.form.get("emotion", "")).lower()
    prompt = (request.form.get("prompt", "") or f"{emotion} abstract scene").strip()

    if emotion not in EMOTIONS:
        log_event("generate_invalid_emotion", user=session["username"], data={"emotion": emotion})
        return render_template("_result_card.html", error="Please choose a valid emotion.")

    image_url = build_image_url(prompt, emotion)
    advice = advice_for(emotion)
    log_event("generate", user=session["username"], data={"emotion": emotion, "prompt": prompt})

    return render_template("_result_card.html",
                           emotion=emotion, prompt=prompt, image_url=image_url, advice=advice)

# -------------------------------
# Image proxy (prevents cert prompts)
# -------------------------------
@bp.get("/api/proxy-image")
@login_required
def proxy_image():
    """
    Proxies a remote image (Pollinations only) through this server to avoid
    client-side TLS/certificate prompts in the browser.
    Usage: /api/proxy-image?url=<encoded-remote-url>
    """
    raw_url = (request.args.get("url") or "").strip()
    if not raw_url:
        return Response("Missing url", status=400, mimetype="text/plain")

    # urlparse never raises for simple strings; validate scheme/host below
    parsed = urlparse(raw_url)
    host = (parsed.hostname or "").lower()

    if parsed.scheme not in ("http", "https") or host not in ALLOWED_IMAGE_HOSTS:
        return Response("Host not allowed", status=400, mimetype="text/plain")

    try:
        resp = requests.get(raw_url, timeout=60)
        resp.raise_for_status()
    except requests.RequestException as e:
        return Response(f"Upstream error: {e}", status=502, mimetype="text/plain")

    content_type = resp.headers.get("Content-Type", "image/png")
    bio = io.BytesIO(resp.content)
    out = send_file(bio, mimetype=content_type)
    out.headers["Cache-Control"] = "no-store"
    return out

# Store feedback (HTMX)
@bp.post("/feedback")
@login_required
def feedback():
    username = session["username"]
    data = {
        "emotion": request.form.get("emotion",""),
        "prompt": request.form.get("prompt",""),
        "image_url": request.form.get("image_url",""),
        "advice": request.form.get("advice",""),
        "predicted_correct": 1 if request.form.get("predicted_correct") == "1" else 0,
        "advice_ok": 1 if request.form.get("advice_ok") == "1" else 0,
        "comments": (request.form.get("comments","") or "").strip(),
        "created_at": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    conn = get_db()
    c = conn.cursor()
    c.execute("""
        INSERT INTO feedback (username, emotion, prompt, image_url, advice, predicted_correct, advice_ok, comments, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (username, data["emotion"], data["prompt"], data["image_url"], data["advice"],
          data["predicted_correct"], data["advice_ok"], data["comments"], data["created_at"]))
    conn.commit()
    conn.close()

    log_event("feedback_submit", user=username, data={
        "emotion": data["emotion"],
        "predicted_correct": data["predicted_correct"],
        "advice_ok": data["advice_ok"]
    })

    return render_template("_result_card.html", saved=True)

# -------------------------------
# Admin dashboard + CSV export
# -------------------------------
@bp.get("/admin")
@login_required
@admin_required
def admin():
    """Legacy admin page - redirect to new dashboard"""
    return redirect(url_for('main.admin_dashboard'))

@bp.get("/export")
@login_required
@admin_required
def export_feedback():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, created_at, username, emotion, prompt, image_url, advice, predicted_correct, advice_ok, comments
        FROM feedback ORDER by id ASC
    """)
    data = cur.fetchall()
    conn.close()

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["id","created_at","username","emotion","prompt","image_url","advice","matched","advice_ok","comments"])
    for row in data:
        w.writerow(row)
    mem = io.BytesIO(out.getvalue().encode("utf-8"))

    log_event("admin_export_csv", user=session["username"], data={"rows": len(data)})
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name="feedback.csv")


@bp.post("/admin/seed-users")
@login_required
@admin_required
def admin_seed_users():
    """
    Create a data/users.json file with some demo users and hashed passwords.
    After this runs once, non-admin login will work with those pairs.
    """
    from werkzeug.security import generate_password_hash

    # where to write users.json (project_root/data/users.json)
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    data_dir = os.path.join(project_root, "data")
    os.makedirs(data_dir, exist_ok=True)
    users_path = os.path.join(data_dir, "users.json")

    # demo users: username -> plain password
    demo = {
        "paul": "paul123",
        "sarah": "sarah123",
        "john": "john123",
        "dana": "dana123",
        "eric": "eric123",
        "fran": "fran123",
        "guest": "guest123",
    }

    # build json list with password_hash
    out = []
    for u, p in demo.items():
        out.append({"username": u, "password_hash": generate_password_hash(p)})

    # write file
    with open(users_path, "w", encoding="utf-8") as fp:
        json.dump(out, fp, ensure_ascii=False, indent=2)

    log_event("admin_seed_users", user=session.get("username"), data={"count": len(out), "path": users_path})
    # back to Admin page
    return redirect(url_for("main.admin_dashboard"))


# -------------------------------
# Admin: Log Viewer + downloads
# -------------------------------
def _read_jsonl(path: str):
    rows = []
    if not os.path.exists(path):
        return rows
    from json import JSONDecodeError
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except JSONDecodeError:
                continue
            rows.append({
                "ts": str(obj.get("ts","")),
                "event": str(obj.get("event","")),
                "user": str(obj.get("user","")),
                "data": json.dumps(obj.get("data",{}), ensure_ascii=False)
            })
    return rows

def _read_csv(path: str):
    rows = []
    if not os.path.exists(path):
        return rows
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append({
                "ts": r.get("ts",""),
                "event": r.get("event",""),
                "user": r.get("user",""),
                "data": r.get("data_json","")
            })
    return rows

def _filter_logs(rows, event_filter, user_filter, start_date, end_date):
    if event_filter:
        rows = [r for r in rows if event_filter in (r.get("event","").lower())]
    if user_filter:
        rows = [r for r in rows if user_filter in (r.get("user","").lower())]
    if start_date or end_date:
        def in_range(ts: str):
            date_only = ts.split(" ", 1)[0] if ts else ""
            if not date_only:
                return False
            if start_date and date_only < start_date:
                return False
            if end_date and date_only > end_date:
                return False
            return True
        rows = [r for r in rows if in_range(r.get("ts",""))]
    return rows

@bp.get("/admin/logs")
@login_required
@admin_required
def admin_logs():
    import math

    page = max(int(request.args.get("page", 1)), 1)
    source = (request.args.get("source","auto") or "auto").lower()
    event_filter = (request.args.get("event","") or "").strip().lower()
    user_filter  = (request.args.get("user","")  or "").strip().lower()
    start_date   = (request.args.get("start","") or "").strip()
    end_date     = (request.args.get("end","")   or "").strip()

    jsonl_path = current_app.config["LOG_JSONL_PATH"]
    csv_path   = current_app.config["LOG_CSV_PATH"]

    if source == "jsonl":
        rows = _read_jsonl(jsonl_path); chosen = "jsonl"
    elif source == "csv":
        rows = _read_csv(csv_path);     chosen = "csv"
    else:
        if os.path.exists(jsonl_path):
            rows = _read_jsonl(jsonl_path); chosen = "jsonl"
        else:
            rows = _read_csv(csv_path);     chosen = "csv"

    rows = list(reversed(rows))
    rows = _filter_logs(rows, event_filter, user_filter, start_date, end_date)

    event_counts = Counter(r["event"] for r in rows)
    per_date     = Counter(r["ts"].split()[0] for r in rows if r["ts"])
    per_user     = Counter(r["user"]          for r in rows if r["user"])

    dates_sorted  = sorted(per_date.keys())
    events_sorted = sorted(event_counts.keys())
    users_sorted  = [u for u, _ in per_user.most_common(10)]

    chart_data = {
        "events": {"labels": events_sorted, "values": [event_counts[e] for e in events_sorted]},
        "dates":  {"labels": dates_sorted,  "values": [per_date[d] for d in dates_sorted]},
        "users":  {"labels": users_sorted,  "values": [per_user[u] for u in users_sorted]},
    }

    page_size = int(current_app.config.get("LOG_VIEW_PAGE_SIZE", 25))
    total = len(rows)
    pages = max(math.ceil(total / page_size), 1)
    start = (page - 1) * page_size
    end   = start + page_size
    page_rows = rows[start:end]

    # Calculate applied filters for the template
    applied_filters = []
    if event_filter:
        applied_filters.append(f"Event: {event_filter}")
    if user_filter:
        applied_filters.append(f"User: {user_filter}")
    if start_date:
        applied_filters.append(f"Start: {start_date}")
    if end_date:
        applied_filters.append(f"End: {end_date}")

    return render_template(
        "logs.html",
        rows=page_rows,
        page=page,
        pages=pages,
        total=total,
        source=chosen,
        event=event_filter,
        user_filter=user_filter,
        start=start_date,
        end=end_date,
        chart_data=chart_data,
        applied_filters=applied_filters,
        F=current_app.jinja_env.filters
    )

@bp.get("/admin/logs/download")
@login_required
@admin_required
def admin_logs_download():
    src = (request.args.get("source","jsonl") or "jsonl").lower()
    path = current_app.config["LOG_JSONL_PATH"] if src == "jsonl" else current_app.config["LOG_CSV_PATH"]

    if not os.path.exists(path):
        return Response("", mimetype="text/plain", headers={
            "Content-Disposition": f'attachment; filename="{os.path.basename(path)}"'
        })
    return send_file(path, as_attachment=True, download_name=os.path.basename(path))

@bp.post("/admin/logs/download-filtered")
@login_required
@admin_required
def admin_logs_download_filtered():
    chosen = (request.args.get("source","auto") or "auto").lower()
    event_filter = (request.form.get("event","") or "").strip().lower()
    user_filter  = (request.form.get("user","")  or "").strip().lower()
    start_date   = (request.form.get("start","") or "").strip()
    end_date     = (request.form.get("end","")   or "").strip()

    jsonl_path = current_app.config["LOG_JSONL_PATH"]
    csv_path   = current_app.config["LOG_CSV_PATH"]

    # Choose source; avoid assigning to 'source' and never reading it
    if chosen == "jsonl":
        rows = _read_jsonl(jsonl_path)
    elif chosen == "csv":
        rows = _read_csv(csv_path)
    else:
        rows = _read_jsonl(jsonl_path) if os.path.exists(jsonl_path) else _read_csv(csv_path)

    rows = _filter_logs(rows, event_filter, user_filter, start_date, end_date)

    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["ts","event","user","data"])
    for r in rows:
        w.writerow([r["ts"], r["event"], r["user"], r["data"]])

    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": 'attachment; filename="filtered_logs.csv"'}
    )

@bp.post("/admin/logs/download-filtered-xlsx")
@login_required
@admin_required
def admin_logs_download_filtered_xlsx():
    # Conditional import for openpyxl
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response(
            "openpyxl package is required for Excel export. Please install it with: pip install openpyxl",
            status=400,
            mimetype="text/plain"
        )

    chosen = (request.args.get("source","auto") or "auto").lower()
    event_filter = (request.form.get("event","") or "").strip().lower()
    user_filter  = (request.form.get("user","")  or "").strip().lower()
    start_date   = (request.form.get("start","") or "").strip()
    end_date     = (request.form.get("end","")   or "").strip()

    jsonl_path = current_app.config["LOG_JSONL_PATH"]
    csv_path   = current_app.config["LOG_CSV_PATH"]

    if chosen == "jsonl":
        rows = _read_jsonl(jsonl_path)
    elif chosen == "csv":
        rows = _read_csv(csv_path)
    else:
        rows = _read_jsonl(jsonl_path) if os.path.exists(jsonl_path) else _read_csv(csv_path)

    rows = _filter_logs(rows, event_filter, user_filter, start_date, end_date)

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Logs"
    ws.append(["Timestamp", "Event", "User", "Data"])
    for r in rows:
        ws.append([r["ts"], r["event"], r["user"], r["data"]])

    for col in ws.columns:
        length = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = length + 2

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="filtered_logs.xlsx"'}
    )

# -------------------------------
# Admin helper: seed sample feedback (optional)
# -------------------------------
@bp.post("/admin/seed-feedback")
@login_required
@admin_required
def admin_seed_feedback():
    import random

    try:
        n = int(request.form.get("n", "25"))
        n = max(1, min(n, 500))
    except ValueError:
        n = 25

    sample_users = ["alice", "bob", "charlie", "dana", "eric", "fran", "guest"]
    sample_prompts = [
        "bright morning after good news",
        "walking in soft rain",
        "stormy night city lights",
        "calm lake at dusk",
        "cozy reading nook",
        "gentle ocean breeze",
        "quiet forest trail",
    ]

    conn = get_db()
    cur = conn.cursor()
    now = dt.datetime.now()

    for _ in range(n):
        emotion = random.choice(EMOTIONS)
        prompt = random.choice(sample_prompts)
        username = random.choice(sample_users)
        advice = advice_for(emotion)
        created_at = (now - dt.timedelta(
            days=random.randint(0, 30),
            hours=random.randint(0, 23),
            minutes=random.randint(0, 59))
        ).strftime("%Y-%m-%d %H:%M:%S")

        poll_prompt = f"{emotion} {prompt}".strip()
        # noinspection SpellCheckingInspection
        image_url = f"https://image.pollinations.ai/prompt/{quote_plus(poll_prompt)}?width=512&height=512&nologo=true"

        cur.execute(
            """
            INSERT INTO feedback
              (username, emotion, prompt, image_url, advice, predicted_correct, advice_ok, comments, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                username,
                emotion,
                prompt,
                image_url,
                advice,
                random.randint(0, 1),
                random.randint(0, 1),
                "",
                created_at,
            ),
        )

    conn.commit()
    conn.close()
    log_event("admin_seed_feedback", user=session.get("username"), data={"n": n})
    return redirect(url_for("main.admin_dashboard"))


# -------------------------------
# New Admin Dashboard with Analytics
# -------------------------------
@bp.get("/admin/dashboard")
@login_required
@admin_required
def admin_dashboard():
    """Admin dashboard with analytics and export options."""
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    emotion_filter = request.args.get('emotion')
    has_feedback = request.args.get('has_feedback')

    # Build query
    query = """
        SELECT f.*, COUNT(fb.id) as feedback_count 
        FROM feedback f 
        LEFT JOIN feedback fb ON f.id = fb.id
    """
    conditions = []
    params = []

    if start_date:
        conditions.append("f.created_at >= ?")
        params.append(start_date)
    if end_date:
        conditions.append("f.created_at <= ?")
        params.append(end_date + " 23:59:59")
    if emotion_filter:
        conditions.append("f.emotion = ?")
        params.append(emotion_filter)
    if has_feedback == "1":
        conditions.append("fb.id IS NOT NULL")
    elif has_feedback == "0":
        conditions.append("fb.id IS NULL")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += " GROUP BY f.id ORDER BY f.created_at DESC LIMIT 100"

    conn = get_db()
    generations = conn.execute(query, params).fetchall()

    # Get statistics
    stats = {
        'total_generations': conn.execute("SELECT COUNT(*) FROM feedback").fetchone()[0],
        'total_users': conn.execute("SELECT COUNT(DISTINCT username) FROM feedback").fetchone()[0],
        'today_generations': conn.execute(
            "SELECT COUNT(*) FROM feedback WHERE DATE(created_at) = DATE('now')"
        ).fetchone()[0],
        'avg_rating': conn.execute(
            "SELECT AVG((predicted_correct + advice_ok) / 2.0) FROM feedback WHERE predicted_correct IS NOT NULL"
        ).fetchone()[0] or 0
    }

    # Get emotion distribution
    emotion_stats = dict(conn.execute(
        "SELECT emotion, COUNT(*) FROM feedback GROUP BY emotion"
    ).fetchall())

    # Get daily statistics
    daily_stats = dict(conn.execute(
        "SELECT DATE(created_at), COUNT(*) FROM feedback GROUP BY DATE(created_at) ORDER BY DATE(created_at) DESC LIMIT 7"
    ).fetchall())

    conn.close()

    return render_template(
        "admin_dashboard.html",
        generations=generations,
        stats=stats,
        emotion_stats=emotion_stats,
        daily_stats=daily_stats,
        emotions=EMOTIONS
    )


@bp.get("/admin/export/<format>")
@login_required
@admin_required
def export_data(format):
    """Export data in various formats."""
    # Get filter parameters
    emotion_filter = request.args.get('emotion')

    # Build query
    if emotion_filter:
        data = get_db().execute(
            "SELECT * FROM feedback WHERE emotion = ? ORDER BY created_at DESC",
            (emotion_filter,)
        ).fetchall()
    else:
        data = get_db().execute(
            "SELECT * FROM feedback ORDER BY created_at DESC"
        ).fetchall()

    # Convert to DataFrame
    df = pd.DataFrame([dict(row) for row in data])

    if format == 'csv':
        output = BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        return send_file(
            output,
            mimetype='text/csv',
            as_attachment=True,
            download_name='mood_app_data.csv'
        )

    elif format == 'excel':
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Generations')

            # Add summary sheet
            summary_data = []
            for emotion in EMOTIONS:
                count = len(df[df['emotion'] == emotion])
                summary_data.append({'Emotion': emotion, 'Count': count})

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, index=False, sheet_name='Summary')

        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='mood_app_data.xlsx'
        )

    elif format == 'pdf':
        # For PDF, we'll create a simple report
        # In a real implementation, you might use ReportLab or WeasyPrint
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        p.drawString(100, 750, "AI Art Mood App - Data Report")
        p.drawString(100, 730, f"Emotion Filter: {emotion_filter or 'All'}")
        p.drawString(100, 710, f"Total Records: {len(df)}")

        y = 690
        for _, row in df.iterrows():
            if y < 100:
                p.showPage()
                y = 750
            p.drawString(100, y, f"{row['created_at']} - {row['emotion']} - {row['username']}")
            y -= 20

        p.save()
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name='mood_app_report.pdf'
        )

    return redirect(url_for('main.admin_dashboard'))